import openpyxl
import random

import openpyxl.chart

filetype = '.xlsx'
filename = 'excel_file/Excel_Sample' + filetype

wb = openpyxl.load_workbook(filename)
sheet = wb['sample']

sum = 0
for row in range(1,11):
  sheet.cell(row = row, column = 1).value = random.uniform(1,10)
  sum += sheet.cell(row = row, column = 1).value
  
sheet['A11'].value = sum


values = openpyxl.chart.Reference(sheet, min_col=1, min_row=1, max_col=1, max_row=10)
chart = openpyxl.chart.LineChart()
chart.add_data(values)
sheet.add_chart(chart, 'A12')

wb.save(filename)
wb.close()